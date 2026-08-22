"""
Админ-панель (API).

Доступ только для пользователей с is_admin=True.
Функции:
  * список грузчиков, блокировка/разблокировка по ID;
  * подтверждение/отклонение оплат (баланс увеличивается при подтверждении);
  * полный CRUD заказов (добавление вручную, редактирование, удаление);
  * управление регионами;
  * настройка комиссии;
  * статистика и журнал действий.
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone

import io

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from ..config import settings
from ..database import get_db
from ..deps import get_current_admin
from ..models import (
    AdminLog,
    Order,
    Payment,
    PromoCode,
    Referral,
    Region,
    Review,
    Setting,
    TakenOrder,
    User,
)
from ..schemas import (
    MessageOut,
    OrderCreate,
    OrderOut,
    PaymentOut,
    PromoCodeCreate,
    PromoCodeOut,
    PromoCodeUpdate,
    RegionCreate,
    SettingsUpdate,
    UserOut,
)
from ..serializers import serialize_order
from ..services import settings_store, telegram
from .orders import get_commission
from .ws import broadcast_orders_update

router = APIRouter(prefix="/admin", tags=["admin"])


def _log(db: Session, admin: User, action: str, details: str) -> None:
    """Короткая запись в журнал действий."""
    db.add(AdminLog(user_id=admin.id, action=action, details=details))


# ============================ ГРУЗЧИКИ ===================================


@router.get("/users", response_model=list[UserOut], summary="Список всех грузчиков")
def list_users(
    search: str | None = None,
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Все пользователи: ID, имя, телефон, баланс, статус."""
    q = select(User).order_by(User.created_at.desc())
    if search:
        like = f"%{search.strip()}%"
        q = q.where(User.phone.ilike(like) | User.name.ilike(like) | User.public_id.ilike(like))
    users = db.scalars(q).all()
    return [UserOut.model_validate(u) for u in users]


@router.post("/users/{user_id}/block", response_model=MessageOut, summary="Заблокировать грузчика")
def block_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Заблокировать грузчика по ID. Заблокированный не может войти и брать заказы."""
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if user.is_admin:
        raise HTTPException(status_code=400, detail="Нельзя заблокировать администратора")

    user.is_blocked = True
    _log(db, admin, "block_user", f"Заблокирован грузчик {user.public_id} ({user.phone})")
    db.commit()
    telegram.notify_user_blocked(user.public_id, True)
    return MessageOut(message=f"Грузчик {user.public_id} заблокирован")


@router.post("/users/{user_id}/unblock", response_model=MessageOut, summary="Разблокировать грузчика")
def unblock_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Снять блокировку с грузчика."""
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    user.is_blocked = False
    _log(db, admin, "unblock_user", f"Разблокирован грузчик {user.public_id}")
    db.commit()
    telegram.notify_user_blocked(user.public_id, False)
    return MessageOut(message=f"Грузчик {user.public_id} разблокирован")


# ============================ ОПЛАТЫ =====================================


@router.get("/payments", response_model=list[PaymentOut], summary="Заявки на пополнение")
def list_payments(
    status_filter: str | None = None,
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Все заявки на пополнение (с именем и ID грузчика)."""
    q = (
        select(Payment)
        .options(joinedload(Payment.user))
        .order_by(Payment.created_at.desc())
    )
    if status_filter:
        q = q.where(Payment.status == status_filter)
    payments = db.scalars(q).all()
    return [
        PaymentOut(
            id=p.id, user_id=p.user_id,
            user_name=p.user.name if p.user else None,
            user_public_id=p.user.public_id if p.user else None,
            amount=p.amount, receipt_file=p.receipt_file, purpose=p.purpose,
            status=p.status, created_at=p.created_at, confirmed_at=p.confirmed_at,
        )
        for p in payments
    ]


def _confirm_payment(db: Session, admin: User, payment_id: int, approve: bool) -> Payment:
    """
    Общая логика подтверждения/отклонения оплаты.
    Используется и HTTP-эндпоинтом, и Telegram-ботом.
    """
    payment = db.get(Payment, payment_id)
    if payment is None:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    if payment.status != "pending":
        raise HTTPException(status_code=400,
                            detail=f"Заявка уже обработана (статус: {payment.status})")

    if approve:
        user = db.get(User, payment.user_id)
        if user is None:
            raise HTTPException(status_code=404, detail="Грузчик не найден")

        # Что именно оплачено, зависит от назначения платежа:
        #  * topup          — пополнение баланса (обычное правило);
        #  * phone_unlock   — разовый доступ к телефонам заказчиков;
        #  * top20          — режим ТОП-20 на сутки.
        purpose = payment.purpose or "topup"
        if purpose == "phone_unlock":
            user.phone_unlocked = True
            action_note = "доступ к телефонам заказчиков открыт"
        elif purpose == "top20":
            now = datetime.now(timezone.utc).replace(tzinfo=None)
            base = user.top20_until
            if base is not None and base.tzinfo is not None:
                base = base.replace(tzinfo=None)
            # Продлеваем от текущего срока, если режим ещё активен
            user.top20_until = max(now, base or now) + timedelta(days=1)
            action_note = "режим ТОП-20 активирован на сутки"
        else:
            user.balance += payment.amount
            action_note = f"баланс пополнен на {payment.amount}₽"

            # Реферальный бонус: пригласивший получает REFERRAL_BONUS один раз,
            # когда приглашённый пополнил баланс на >= REFERRAL_TOPUP_MIN.
            if payment.amount >= settings.REFERRAL_TOPUP_MIN:
                ref = db.scalar(
                    select(Referral).where(Referral.referred_id == user.id)
                )
                if ref is not None and not ref.bonus_paid:
                    referrer = db.get(User, ref.referrer_id)
                    if referrer is not None:
                        referrer.balance = (referrer.balance or 0) + settings.REFERRAL_BONUS
                        ref.bonus_amount = settings.REFERRAL_BONUS
                        ref.bonus_paid = True
                        action_note += (f"; пригласившему {referrer.public_id} "
                                        f"начислен реферальный бонус "
                                        f"{settings.REFERRAL_BONUS}₽")
                        _log(db, admin, "referral_bonus",
                             f"Бонус за приглашение {user.public_id}: "
                             f"+{settings.REFERRAL_BONUS}₽ (пополнение "
                             f"{payment.amount}₽)")

        payment.status = "confirmed"
        payment.confirmed_at = datetime.now(timezone.utc)
        payment.confirmed_by = admin.id
        details = (f"Подтверждена оплата #{payment.id} ({purpose}): {action_note} "
                   f"— грузчик {user.public_id}")
    else:
        payment.status = "rejected"
        payment.confirmed_at = datetime.now(timezone.utc)
        payment.confirmed_by = admin.id
        details = f"Отклонена оплата #{payment.id}"

    _log(db, admin, "confirm_payment" if approve else "reject_payment", details)
    db.commit()
    db.refresh(payment)
    return payment


@router.post("/payments/{payment_id}/confirm", response_model=PaymentOut,
             summary="Подтвердить оплату")
def confirm_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Подтвердить оплату: баланс грузчика увеличивается на сумму."""
    payment = _confirm_payment(db, admin, payment_id, approve=True)
    return PaymentOut(
        id=payment.id, user_id=payment.user_id, amount=payment.amount,
        receipt_file=payment.receipt_file, purpose=payment.purpose,
        status=payment.status, created_at=payment.created_at,
        confirmed_at=payment.confirmed_at,
        user_public_id=payment.user.public_id if payment.user else None,
        user_name=payment.user.name if payment.user else None,
    )


@router.post("/payments/{payment_id}/reject", response_model=PaymentOut,
             summary="Отклонить оплату")
def reject_payment(
    payment_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Отклонить заявку на пополнение (баланс не меняется)."""
    payment = _confirm_payment(db, admin, payment_id, approve=False)
    return PaymentOut(
        id=payment.id, user_id=payment.user_id, amount=payment.amount,
        receipt_file=payment.receipt_file, purpose=payment.purpose,
        status=payment.status, created_at=payment.created_at,
        confirmed_at=payment.confirmed_at,
        user_public_id=payment.user.public_id if payment.user else None,
        user_name=payment.user.name if payment.user else None,
    )


# ============================ ЗАКАЗЫ =====================================


def _resolve_region(db: Session, region_name: str) -> Region:
    """Найти регион по имени; если нет — создать (админ добавляет города сам)."""
    region = db.scalar(select(Region).where(Region.name == region_name.strip()))
    if region is None:
        region = Region(name=region_name.strip())
        db.add(region)
        db.flush()
    return region


@router.get("/orders", response_model=list[OrderOut], summary="Все заказы (админ)")
def admin_orders(
    date_filter: str | None = Query("today", description="today|all"),
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Список заказов для админа: сегодня или все. Телефоны видны всегда."""
    q = select(Order).options(
        joinedload(Order.region),
        joinedload(Order.taken).joinedload(TakenOrder.user),
    ).order_by(Order.published_at.desc())
    if date_filter == "today":
        now = datetime.now(timezone.utc)
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        q = q.where(Order.published_at >= start)
    orders = db.scalars(q).all()
    # Админ всегда видит телефоны: передаём огромный баланс
    return [serialize_order(o, balance=10**9, taken=o.taken) for o in orders]


@router.post("/orders", response_model=OrderOut, summary="Добавить заказ вручную")
async def admin_create_order(
    data: OrderCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Создать заказ (например, поступивший по телефону). Публикуется на сегодня."""
    region = _resolve_region(db, data.region_name)
    order = Order(
        region_id=region.id,
        street=data.street.strip(),
        house=data.house.strip(),
        apartment=data.apartment,
        entrance=data.entrance,
        floor=data.floor,
        landmarks=data.landmarks,
        phone=data.phone,
        price=data.price,
        hourly_rate=data.hourly_rate,
        weight=data.weight,
        deadline=data.deadline,
        duration_min=data.duration_min,
        duration_max=data.duration_max,
        category=data.category,
        urgency=data.urgency,
        description=data.description,
        latitude=data.latitude,
        longitude=data.longitude,
        status="active",
        source="admin",
    )
    db.add(order)
    db.flush()
    _log(db, admin, "create_order", f"Создан заказ #{order.id} в регионе {region.name}")
    db.commit()
    db.refresh(order)

    await broadcast_orders_update(region=region.name)
    return OrderOut(**serialize_order(order, balance=10**9))


@router.patch("/orders/{order_id}", response_model=OrderOut, summary="Редактировать заказ")
async def admin_update_order(
    order_id: int,
    data: OrderCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Обновить поля существующего заказа."""
    order = db.get(Order, order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Заказ не найден")

    region = _resolve_region(db, data.region_name)
    order.region_id = region.id
    order.street = data.street.strip()
    order.house = data.house.strip()
    order.apartment = data.apartment
    order.entrance = data.entrance
    order.floor = data.floor
    order.landmarks = data.landmarks
    order.phone = data.phone
    order.price = data.price
    order.hourly_rate = data.hourly_rate
    order.weight = data.weight
    order.deadline = data.deadline
    order.duration_min = data.duration_min
    order.duration_max = data.duration_max
    order.category = data.category
    order.urgency = data.urgency
    order.description = data.description
    order.latitude = data.latitude
    order.longitude = data.longitude

    _log(db, admin, "update_order", f"Изменён заказ #{order.id}")
    db.commit()
    db.refresh(order)

    await broadcast_orders_update(region=region.name)
    return OrderOut(**serialize_order(order, balance=10**9))


@router.delete("/orders/{order_id}", response_model=MessageOut, summary="Удалить заказ")
async def admin_delete_order(
    order_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Удалить заказ из системы."""
    order = db.get(Order, order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Заказ не найден")

    _log(db, admin, "delete_order", f"Удалён заказ #{order.id}")
    db.delete(order)
    db.commit()

    await broadcast_orders_update()
    return MessageOut(message=f"Заказ #{order_id} удалён")


@router.post("/orders/{order_id}/complete", response_model=MessageOut,
             summary="Отметить заказ выполненным")
def admin_complete_order(
    order_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Пометить взятый заказ как выполненный."""
    taken = db.scalar(select(TakenOrder).where(TakenOrder.order_id == order_id))
    if taken is None:
        raise HTTPException(status_code=400, detail="Заказ ещё никем не взят")

    taken.completed_at = datetime.now(timezone.utc)
    taken.order.status = "completed"
    _log(db, admin, "complete_order", f"Заказ #{order_id} выполнен")
    db.commit()
    return MessageOut(message=f"Заказ #{order_id} отмечен выполненным")


# ============================ РЕГИОНЫ ====================================


@router.get("/regions", response_model=list[dict], summary="Список регионов")
def admin_regions(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Все регионы с количеством заказов."""
    regions = db.scalars(select(Region).order_by(Region.name)).all()
    result = []
    for r in regions:
        count = db.scalar(select(func.count(Order.id)).where(Order.region_id == r.id)) or 0
        result.append({"id": r.id, "name": r.name, "is_active": r.is_active, "orders_count": count})
    return result


@router.post("/regions", response_model=dict, summary="Добавить регион")
def admin_create_region(
    data: RegionCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Добавить новый город/регион."""
    existing = db.scalar(select(Region).where(Region.name == data.name.strip()))
    if existing:
        raise HTTPException(status_code=409, detail="Регион уже существует")
    region = Region(name=data.name.strip())
    db.add(region)
    _log(db, admin, "create_region", f"Добавлен регион {region.name}")
    db.commit()
    return {"id": region.id, "name": region.name, "is_active": True}


@router.delete("/regions/{region_id}", response_model=MessageOut, summary="Удалить регион")
def admin_delete_region(
    region_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Удалить регион (заказы региона удаляются вместе с ним)."""
    region = db.get(Region, region_id)
    if region is None:
        raise HTTPException(status_code=404, detail="Регион не найден")
    _log(db, admin, "delete_region", f"Удалён регион {region.name}")
    db.delete(region)
    db.commit()
    return MessageOut(message=f"Регион «{region.name}» удалён")


# ============================ НАСТРОЙКИ ==================================


@router.get("/settings", response_model=dict, summary="Текущие настройки")
def admin_settings(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Вернуть настройки (комиссия, цены услуг, порог баланса, реквизиты)."""
    return {
        "commission": get_commission(db),
        "min_topup": settings.MIN_TOPUP_AMOUNT,
        "phone_visible_balance": settings.PHONE_VISIBLE_BALANCE,
        "phone_unlock_amount": settings_store.get_int_setting(
            db, "phone_unlock_amount", settings.PHONE_UNLOCK_AMOUNT
        ),
        "top20_price": settings_store.get_int_setting(
            db, "top20_price", settings.TOP20_DAILY_PRICE
        ),
        "bank": {
            "name": settings.BANK_NAME,
            "phone": settings.BANK_PHONE,
            "card": settings.BANK_CARD,
            "holder": settings.BANK_HOLDER,
        },
    }


@router.put("/settings", response_model=dict, summary="Обновить настройки")
def admin_update_settings(
    data: SettingsUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Изменить размер комиссии и цены услуг (телефоны, ТОП-20)."""
    changed = []
    if data.commission is not None:
        setting = db.get(Setting, "commission")
        if setting is None:
            setting = Setting(key="commission", value=str(data.commission))
            db.add(setting)
        else:
            setting.value = str(data.commission)
        changed.append(f"комиссия — {data.commission}₽")
    if data.phone_unlock_amount is not None:
        settings_store.set_int_setting(db, "phone_unlock_amount", data.phone_unlock_amount)
        changed.append(f"доступ к телефонам — {data.phone_unlock_amount}₽")
    if data.top20_price is not None:
        settings_store.set_int_setting(db, "top20_price", data.top20_price)
        changed.append(f"ТОП-20 за сутки — {data.top20_price}₽")
    if changed:
        _log(db, admin, "update_settings", "Изменены настройки: " + ", ".join(changed))
        db.commit()
    return {
        "commission": get_commission(db),
        "phone_unlock_amount": settings_store.get_int_setting(
            db, "phone_unlock_amount", settings.PHONE_UNLOCK_AMOUNT
        ),
        "top20_price": settings_store.get_int_setting(
            db, "top20_price", settings.TOP20_DAILY_PRICE
        ),
    }


# ============================ СТАТИСТИКА =================================


def _day_key(dt: datetime | None) -> str:
    """Дата в формате YYYY-MM-DD (с учётом naive/aware datetime из БД)."""
    if dt is None:
        return ""
    return (dt.replace(tzinfo=None) if dt.tzinfo else dt).date().isoformat()


@router.get("/stats", summary="Статистика платформы (с графиками)")
def admin_stats(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """
    Расширенная статистика для админ-панели:
      * пользователи (новые за день/неделю/месяц, активные за 30 дней);
      * заказы (всего/взято/выполнено, по источникам, конверсия);
      * финансы (доход от комиссий, подтверждённые пополнения, средний чек);
      * графики за 14 дней: доход, публикации/взятия/выполнения, новые грузчики;
      * распределение заказов по категориям и отзывы.
    """
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = now - timedelta(weeks=1)
    month_start = now - timedelta(days=30)

    # --- Базовые счётчики ---
    total_users = db.scalar(select(func.count(User.id))) or 0
    blocked_users = db.scalar(
        select(func.count(User.id)).where(User.is_blocked.is_(True))
    ) or 0
    new_today = db.scalar(
        select(func.count(User.id)).where(User.created_at >= today_start)
    ) or 0
    new_week = db.scalar(
        select(func.count(User.id)).where(User.created_at >= week_start)
    ) or 0
    new_month = db.scalar(
        select(func.count(User.id)).where(User.created_at >= month_start)
    ) or 0
    active_30d = db.scalar(
        select(func.count(func.distinct(TakenOrder.user_id)))
        .where(TakenOrder.taken_at >= month_start)
    ) or 0

    total_orders = db.scalar(select(func.count(Order.id))) or 0
    today_orders = db.scalar(
        select(func.count(Order.id)).where(Order.published_at >= today_start)
    ) or 0
    week_orders = db.scalar(
        select(func.count(Order.id)).where(Order.published_at >= week_start)
    ) or 0
    month_orders = db.scalar(
        select(func.count(Order.id)).where(Order.published_at >= month_start)
    ) or 0

    taken_total = db.scalar(select(func.count(TakenOrder.id))) or 0
    completed_total = db.scalar(
        select(func.count(TakenOrder.id)).where(TakenOrder.completed_at.is_not(None))
    ) or 0

    form_orders = db.scalar(
        select(func.count(Order.id)).where(Order.source == "form")
    ) or 0

    commission_income = db.scalar(
        select(func.coalesce(func.sum(TakenOrder.commission), 0))
    ) or 0
    pending_payments = db.scalar(
        select(func.count(Payment.id)).where(Payment.status == "pending")
    ) or 0
    confirmed_sum = db.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.status == "confirmed")
    ) or 0
    avg_order_price = db.scalar(
        select(func.coalesce(func.avg(Order.price), 0))
    ) or 0

    # --- Графики за последние 14 дней (агрегация в Python — не зависит от БД) ---
    days = [today_start - timedelta(days=i) for i in range(13, -1, -1)]
    day_keys = [d.date().isoformat() for d in days]
    income_by_day: dict[str, int] = {k: 0 for k in day_keys}
    topups_by_day: dict[str, int] = {k: 0 for k in day_keys}
    published_by_day: dict[str, int] = {k: 0 for k in day_keys}
    taken_by_day: dict[str, int] = {k: 0 for k in day_keys}
    completed_by_day: dict[str, int] = {k: 0 for k in day_keys}
    new_users_by_day: dict[str, int] = {k: 0 for k in day_keys}
    orders_by_category: dict[str, int] = {}

    for t in db.scalars(select(TakenOrder)).all():
        k = _day_key(t.taken_at)
        if k in income_by_day:
            income_by_day[k] += t.commission or 0
            taken_by_day[k] += 1
        if t.completed_at and _day_key(t.completed_at) in completed_by_day:
            completed_by_day[_day_key(t.completed_at)] += 1

    for o in db.scalars(select(Order)).all():
        k = _day_key(o.published_at)
        if k in published_by_day:
            published_by_day[k] += 1
        cat = o.category or "прочее"
        orders_by_category[cat] = orders_by_category.get(cat, 0) + 1

    for p in db.scalars(select(Payment).where(Payment.status == "confirmed")).all():
        k = _day_key(p.confirmed_at or p.created_at)
        if k in topups_by_day:
            topups_by_day[k] += p.amount or 0

    for u in db.scalars(select(User)).all():
        k = _day_key(u.created_at)
        if k in new_users_by_day:
            new_users_by_day[k] += 1

    # --- Отзывы ---
    reviews_total = db.scalar(select(func.count(Review.id))) or 0
    reviews_avg = db.scalar(
        select(func.avg(Review.rating)).where(Review.from_role == "customer")
    )
    loader_reviews = db.scalar(
        select(func.count(Review.id)).where(Review.from_role == "loader")
    ) or 0

    # --- Конверсия: опубликовано → взято → выполнено ---
    taken_pct = round(taken_total / total_orders * 100, 1) if total_orders else 0.0
    completed_pct = round(completed_total / total_orders * 100, 1) if total_orders else 0.0

    return {
        "users": {
            "total": total_users,
            "blocked": blocked_users,
            "new_today": new_today,
            "new_week": new_week,
            "new_month": new_month,
            "active_30d": active_30d,
        },
        "orders": {
            "today": today_orders,
            "week": week_orders,
            "month": month_orders,
            "total": total_orders,
            "taken": taken_total,
            "completed": completed_total,
            "by_source": {"admin": total_orders - form_orders, "form": form_orders},
            "conversion": {"taken_pct": taken_pct, "completed_pct": completed_pct},
        },
        "finance": {
            "commission_income": int(commission_income),
            "taken_orders": taken_total,
            "confirmed_topups_sum": int(confirmed_sum),
            "pending_payments": pending_payments,
            "avg_order_price": round(float(avg_order_price), 2),
        },
        "charts": {
            "income_14d": [
                {"date": k, "commission": income_by_day[k], "topups": topups_by_day[k]}
                for k in day_keys
            ],
            "orders_14d": [
                {
                    "date": k,
                    "published": published_by_day[k],
                    "taken": taken_by_day[k],
                    "completed": completed_by_day[k],
                }
                for k in day_keys
            ],
            "new_users_14d": [
                {"date": k, "count": new_users_by_day[k]} for k in day_keys
            ],
            "orders_by_category": [
                {"category": c, "count": n}
                for c, n in sorted(orders_by_category.items(), key=lambda x: -x[1])
            ],
        },
        "reviews": {
            "total": reviews_total,
            "avg_rating": round(float(reviews_avg), 2) if reviews_avg is not None else None,
            "loader_reviews": loader_reviews,
        },
    }


@router.get("/stats/export", summary="Выгрузка статистики в Excel")
def admin_stats_export(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Сформировать .xlsx-файл со всеми данными для отчётности.

    Листы: Статистика, Заказы, Платежи, Грузчики, Отзывы, Промокоды.
    Файл отдаётся как вложение (Content-Disposition: attachment).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl не установлен — добавьте его в requirements.txt",
        )

    wb = Workbook()
    bold = Font(bold=True)

    # --- Лист «Статистика» ---
    ws = wb.active
    ws.title = "Статистика"
    ws.append(["Показатель", "Значение"])
    stats = admin_stats(db, _admin)
    flat = [
        ("Всего грузчиков", stats["users"]["total"]),
        ("Заблокировано", stats["users"]["blocked"]),
        ("Новых за 30 дней", stats["users"]["new_month"]),
        ("Активных за 30 дней", stats["users"]["active_30d"]),
        ("Заказов всего", stats["orders"]["total"]),
        ("Заказов сегодня", stats["orders"]["today"]),
        ("Взято заказов", stats["orders"]["taken"]),
        ("Выполнено заказов", stats["orders"]["completed"]),
        ("Конверсия в взятие, %", stats["orders"]["conversion"]["taken_pct"]),
        ("Конверсия в выполнение, %", stats["orders"]["conversion"]["completed_pct"]),
        ("Доход от комиссий, ₽", stats["finance"]["commission_income"]),
        ("Подтверждено пополнений, ₽", stats["finance"]["confirmed_topups_sum"]),
        ("Ожидают подтверждения", stats["finance"]["pending_payments"]),
        ("Средний чек заказа, ₽", stats["finance"]["avg_order_price"]),
        ("Отзывов всего", stats["reviews"]["total"]),
        ("Средняя оценка заказчиков", stats["reviews"]["avg_rating"]),
    ]
    for row in flat:
        ws.append(row)
    for cell in ws[1]:
        cell.font = bold
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 24

    # --- Лист «Заказы» ---
    ws2 = wb.create_sheet("Заказы")
    ws2.append(["ID", "Регион", "Адрес", "Цена, ₽", "Категория", "Статус", "Источник", "Опубликован"])
    for o in db.scalars(select(Order).order_by(Order.id.desc())).all():
        ws2.append([
            o.id, o.region.name if o.region else "",
            f"{o.street} {o.house}".strip(),
            o.price, o.category, o.status, o.source,
            (o.published_at.replace(tzinfo=None) if o.published_at.tzinfo else o.published_at),
        ])
    for cell in ws2[1]:
        cell.font = bold

    # --- Лист «Платежи» ---
    ws3 = wb.create_sheet("Платежи")
    ws3.append(["ID", "Грузчик", "Сумма, ₽", "Назначение", "Статус", "Создан", "Подтверждён"])
    for p in db.scalars(select(Payment).order_by(Payment.id.desc())).all():
        user = db.get(User, p.user_id)
        ws3.append([
            p.id, user.name if user else "", p.amount, p.purpose, p.status,
            (p.created_at.replace(tzinfo=None) if p.created_at.tzinfo else p.created_at),
            (p.confirmed_at.replace(tzinfo=None) if p.confirmed_at and p.confirmed_at.tzinfo else p.confirmed_at),
        ])
    for cell in ws3[1]:
        cell.font = bold

    # --- Лист «Грузчики» ---
    ws4 = wb.create_sheet("Грузчики")
    ws4.append(["ID", "Публичный ID", "Имя", "Телефон", "Баланс, ₽", "Заблокирован", "Админ", "Регистрация"])
    for u in db.scalars(select(User).order_by(User.id)).all():
        ws4.append([
            u.id, u.public_id, u.name, u.phone, u.balance,
            "да" if u.is_blocked else "нет", "да" if u.is_admin else "нет",
            (u.created_at.replace(tzinfo=None) if u.created_at.tzinfo else u.created_at),
        ])
    for cell in ws4[1]:
        cell.font = bold

    # --- Лист «Отзывы» ---
    ws5 = wb.create_sheet("Отзывы")
    ws5.append(["ID", "Заказ", "Кто", "Роль", "Оценка", "Комментарий", "Дата"])
    for r in db.scalars(select(Review).order_by(Review.id.desc())).all():
        ws5.append([
            r.id, r.order_id,
            (r.from_user.name if r.from_user else (r.from_phone or "")),
            "заказчик" if r.from_role == "customer" else "грузчик",
            r.rating, r.comment or "",
            (r.created_at.replace(tzinfo=None) if r.created_at.tzinfo else r.created_at),
        ])
    for cell in ws5[1]:
        cell.font = bold

    # --- Лист «Промокоды» ---
    ws6 = wb.create_sheet("Промокоды")
    ws6.append(["ID", "Код", "Бонус, ₽", "Лимит", "Активаций", "Активен", "Создан"])
    for pc in db.scalars(select(PromoCode).order_by(PromoCode.id)).all():
        ws6.append([
            pc.id, pc.code, pc.bonus, pc.max_uses, pc.uses_count,
            "да" if pc.is_active else "нет",
            (pc.created_at.replace(tzinfo=None) if pc.created_at.tzinfo else pc.created_at),
        ])
    for cell in ws6[1]:
        cell.font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"stats_{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ============================ ПРОМОКОДЫ ==================================


@router.get("/promocodes", response_model=list[PromoCodeOut], summary="Список промокодов")
def admin_list_promocodes(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Все промокоды с количеством активаций (сначала новые)."""
    return db.scalars(select(PromoCode).order_by(PromoCode.id.desc())).all()


@router.post("/promocodes", response_model=PromoCodeOut, summary="Создать промокод")
def admin_create_promocode(
    data: PromoCodeCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Создать промокод: код (регистр не важен) + бонус + лимит активаций."""
    code = data.code.strip().upper()
    existing = db.scalar(select(PromoCode).where(PromoCode.code == code))
    if existing is not None:
        raise HTTPException(status_code=409, detail="Промокод с таким кодом уже существует")
    promo = PromoCode(code=code, bonus=data.bonus, max_uses=data.max_uses)
    db.add(promo)
    _log(db, admin, "create_promo", f"Промокод {code}: +{data.bonus}₽, лимит {data.max_uses}")
    db.commit()
    db.refresh(promo)
    return promo


@router.patch("/promocodes/{promo_id}", response_model=PromoCodeOut, summary="Изменить промокод")
def admin_update_promocode(
    promo_id: int,
    data: PromoCodeUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Изменить бонус, лимит или включить/выключить промокод."""
    promo = db.get(PromoCode, promo_id)
    if promo is None:
        raise HTTPException(status_code=404, detail="Промокод не найден")
    if data.bonus is not None:
        promo.bonus = data.bonus
    if data.max_uses is not None:
        promo.max_uses = data.max_uses
    if data.is_active is not None:
        promo.is_active = data.is_active
    _log(db, admin, "update_promo", f"Промокод {promo.code} обновлён")
    db.commit()
    db.refresh(promo)
    return promo


@router.delete("/promocodes/{promo_id}", response_model=MessageOut, summary="Удалить промокод")
def admin_delete_promocode(
    promo_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin),
):
    """Удалить промокод."""
    promo = db.get(PromoCode, promo_id)
    if promo is None:
        raise HTTPException(status_code=404, detail="Промокод не найден")
    code = promo.code
    db.delete(promo)
    _log(db, admin, "delete_promo", f"Удалён промокод {code}")
    db.commit()
    return MessageOut(message=f"Промокод {code} удалён")


@router.get("/logs", summary="Журнал действий")
def admin_logs(
    limit: int = 100,
    db: Session = Depends(get_db),
    _admin: User = Depends(get_current_admin),
):
    """Последние записи журнала действий (кто взял заказ, когда и т.д.)."""
    logs = db.scalars(
        select(AdminLog).order_by(AdminLog.created_at.desc()).limit(min(limit, 500))
    ).all()
    return [
        {"id": l.id, "user_id": l.user_id, "action": l.action,
         "details": l.details, "created_at": l.created_at.isoformat()}
        for l in logs
    ]
